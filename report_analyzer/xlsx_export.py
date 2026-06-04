"""Excel export in the customer "TR Summary" layout.

Produces a workbook with two sheets that mirror the supplied template:

* **FX Temp.** - fixed / non-operable products (windows)
* **SD Temp.** - operable products (sliding / patio doors), with extra
  ``Leaf Size`` and split primary/secondary leaf glass columns.

Each test specimen becomes one row, routed to the sheet that matches its
product type. ``write_xlsx`` builds a fresh, styled workbook; ``fill_template``
appends rows into a copy of an existing template file, preserving its styling.
"""

from __future__ import annotations

import datetime as _dt
import re
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from .models import Report, Specimen

# Sheet titles and headers exactly as in the template.
FX_SHEET = "FX Temp."
SD_SHEET = "SD Temp."

FX_HEADERS = [
    "Test Report #", "Report Date", "Specimen #", "DP (psf)", "Test Standard",
    "Frame Size", "D.L.O. Size", "Description", "Glass\n(Ext. to Int.)", "Misc. Note",
]
SD_HEADERS = [
    "Test Report #", "Report Date", "Specimen #", "DP (psf)", "Test Standard",
    "Frame Size", "Leaf Size", "D.L.O. Size", "Description",
    "Glass - Primary Leaf\n(Ext. to Int.)", "Glass - Secondary Leaf\n(Ext. to Int.)",
    "Misc. Note",
]

FX_WIDTHS = [17.9, 11.6, 20.1, 10.4, 26.9, 12.7, 16.1, 19.9, 34.6, 30.6]
SD_WIDTHS = [18.9, 11.6, 11.1, 8.0, 26.9, 12.7, 12.7, 24.3, 19.9, 34.6, 34.6, 30.6]

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FONT = Font(name="Calibri", size=11)
_DATE_FMT = "mm/dd/yyyy"


# --------------------------------------------------------------------------- #
# Field mapping
# --------------------------------------------------------------------------- #


def is_door(report: Report) -> bool:
    """Route doors / sliding / patio products to the SD sheet.

    Honours an explicit ``product_category`` (set by AI extraction) when present;
    otherwise falls back to keyword matching on the product type / model.
    """
    category = (getattr(report, "product_category", "") or "").strip().lower()
    if category in ("door", "window"):
        return category == "door"
    p = (report.product_type + " " + report.series_model).lower()
    return any(word in p for word in ("door", "sliding", "patio", "slider"))


def _parse_date(report: Report):
    """Best effective report date as a real date (revised wins, like the
    template); falls back to the raw string if it cannot be parsed."""
    raw = report.revised_date or report.report_date or report.issue_date
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(raw.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return raw


def _size_compact(size: str) -> str:
    # "216 in x 120 in" -> "216 x 120"
    return re.sub(r"\s*in\b", "", size).strip()


def _dp(spec: Specimen) -> str:
    """Design pressure for the psf column.

    Keeps an already-compact '+X / -Y psf' value; otherwise extracts the psf
    magnitude (e.g. from '+/-2400 Pa (+/-50.13 psf)') as '+50.13 / -50.13 psf'.
    """
    dp = spec.design_pressure
    if re.search(r"[+\-]\s*\d[\d.]*\s*/\s*[+\-]?\s*\d", dp):
        return dp
    m = re.search(r"([\d.]+)\s*psf", dp)
    if m:
        return f"+{m.group(1)} / -{m.group(1)} psf"
    return dp


def _standards(report: Report) -> str:
    return "\n".join(report.test_standards)


def _glass(spec: Specimen) -> str:
    """Glass make-up, one layer per line (Ext. to Int.)."""
    g = spec.construction.glazing
    text = (g.makeup or g.glass_type) if g else ""
    # Break before each layer that begins with a measurement (e.g. 1/8", 0.090").
    return re.sub(r"(?<=\S)\s+(?=\d[\d/.]*\"\s+[\(A-Za-z])", "\n", text)


def _specimen_name(spec: Specimen) -> str:
    """Model/label without trailing dimensions, upper-cased (FX style)."""
    name = spec.label or spec.model or f"Specimen {spec.specimen_id}"
    name = re.split(r"\s+-\s+\d", name)[0]  # drop trailing " - 216\" X 120\""
    return name.strip().upper()


def _description(report: Report, spec: Specimen) -> str:
    glass = spec.construction.glazing
    glass_type = glass.glass_type if glass else ""
    if glass_type and report.product_type:
        return f"{glass_type} ({report.product_type})"
    return glass_type or report.product_type


def _misc_fx(spec: Specimen) -> str:
    g = spec.construction.glazing
    return g.method if g else ""


def _misc_sd(spec: Specimen) -> str:
    return "\n".join(p for p in (spec.label, spec.model) if p)


def _fx_row(report: Report, spec: Specimen) -> list:
    return [
        report.report_number, _parse_date(report), _specimen_name(spec),
        _dp(spec), _standards(report), _size_compact(spec.overall_size),
        spec.daylight_opening or "-", _description(report, spec),
        _glass(spec), _misc_fx(spec),
    ]


def _sd_row(report: Report, spec: Specimen) -> list:
    return [
        report.report_number, _parse_date(report), spec.specimen_id or _specimen_name(spec),
        _dp(spec), _standards(report), _size_compact(spec.overall_size),
        _size_compact(spec.leaf_size) or "-", spec.daylight_opening or "-",
        report.product_type or _description(report, spec),
        _glass(spec), "-", _misc_sd(spec),
    ]


# --------------------------------------------------------------------------- #
# Styling helpers
# --------------------------------------------------------------------------- #


def _style_cell(cell, date: bool = False) -> None:
    cell.font = _FONT
    cell.alignment = _CENTER
    cell.border = _BORDER
    if date and isinstance(cell.value, (_dt.date, _dt.datetime)):
        cell.number_format = _DATE_FMT


def _row_height(values: list) -> float:
    lines = max((str(v).count("\n") + 1) for v in values) if values else 1
    return max(15.0, 15.0 * lines)


def _write_sheet(ws, headers: List[str], widths: List[float], rows: List[list]) -> None:
    ws.append(headers)
    ws.row_dimensions[1].height = 30
    for i, cell in enumerate(ws[1]):
        _style_cell(cell)
        ws.column_dimensions[cell.column_letter].width = widths[i]
    ws.freeze_panes = "A2"
    for r, values in enumerate(rows, start=2):
        ws.append(values)
        ws.row_dimensions[r].height = _row_height(values)
        for cell in ws[r]:
            _style_cell(cell, date=(cell.column == 2))


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #


def split_rows(reports: Iterable[Report]):
    """Return (fx_rows, sd_rows) routed by product type, one row per specimen."""
    fx, sd = [], []
    for report in reports:
        for spec in (report.specimens or [Specimen()]):
            if is_door(report):
                sd.append(_sd_row(report, spec))
            else:
                fx.append(_fx_row(report, spec))
    return fx, sd


def write_xlsx(reports: Iterable[Report], path: str) -> tuple:
    """Write a fresh, styled workbook in the FX/SD template layout."""
    reports = list(reports)
    fx, sd = split_rows(reports)
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb.create_sheet(FX_SHEET), FX_HEADERS, FX_WIDTHS, fx)
    _write_sheet(wb.create_sheet(SD_SHEET), SD_HEADERS, SD_WIDTHS, sd)
    wb.save(path)
    return len(fx), len(sd)


def fill_template(reports: Iterable[Report], template_path: str, out_path: str) -> tuple:
    """Append rows into a copy of an existing template workbook.

    Rows are added below the last used row of the matching sheet; the
    template's own styling, column widths and header rows are left untouched.
    A data-row style is inferred and applied to the new cells.
    """
    reports = list(reports)
    fx, sd = split_rows(reports)
    wb = load_workbook(template_path)
    titles = {t.lower(): t for t in wb.sheetnames}

    def target(*aliases):
        for a in aliases:
            if a.lower() in titles:
                return wb[titles[a.lower()]]
        return None

    for ws, rows in ((target(FX_SHEET, "FX"), fx), (target(SD_SHEET, "SD"), sd)):
        if ws is None or not rows:
            continue
        start = ws.max_row + 1
        for offset, values in enumerate(rows):
            r = start + offset
            for c, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                _style_cell(cell, date=(c == 2))
            ws.row_dimensions[r].height = _row_height(values)
    wb.save(out_path)
    return len(fx), len(sd)
