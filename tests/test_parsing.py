"""Unit tests for the report analyzer.

These tests run entirely on synthetic text and synthetic tables that mimic the
two real report layouts, so they need neither the example PDFs nor PyMuPDF.
They are written to run under pytest *or* as a plain script::

    python -m pytest tests/ -q
    python tests/test_parsing.py
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from report_analyzer.extraction import Document, Page
from report_analyzer.models import CSV_COLUMNS, Report, Specimen
from report_analyzer import parsing
from report_analyzer.parsing import (
    FAMILY_NAFS,
    FAMILY_TAS,
    analyze,
    detect_family,
    labeled_longest,
    _finalize_raw,
    _frame_from_table,
    _glazing_from_table,
    _hardware_from_table,
    _is_header_row,
    _row_to_three,
)


# --------------------------------------------------------------------------- #
# Synthetic fixtures
# --------------------------------------------------------------------------- #

NAFS_PAGE1 = """Report No.:  99001.01-100-50-R1
Report Date:  03/04/2026
TEST REPORT
AAMA/WDMA/CSA 101/I.S.2/A440-17
AAMA/WDMA/CSA 101/I.S.2/A440-22
REPORT NO.:
99001.01-100-50-R1
RENDERED TO:
ACME FENESTRATION
PRODUCT TYPE:
Vinyl Single Hung Window, Type H
SERIES / MODEL: 300-SH / 350-SH
Test Specimen #1
New Construction Frame
300-SH
Summary of Results
Primary Product Designator
Class R - PG40 920 x 1520 (36 x 60)-H
Design Pressure
+/-1920 Pa (+/-40.10 psf)
Air Infiltration @ 1.57 psf
0.5 L/s/m2 (0.10 cfm/ft2)
Air Exfiltration @ 1.57 psf
0.4 L/s/m2 (0.08 cfm/ft2)
Water Penetration Resistance Test Pressure
290 Pa (6.05 psf)
Test Specimen #2
Replacement Frame
350-SH
Summary of Results
Primary Product Designator
Class R - PG40 920 x 1520 (36 x 60)-H
Design Pressure
+/-1440 Pa (+/-30.08 psf)
Air Infiltration @ 1.57 psf
See Specimen #1
Air Exfiltration @ 1.57 psf
See Specimen #1
Water Penetration Resistance Test Pressure
See Specimen #1
Test Completion Date:
02/10/2026
"""

NAFS_PAGE2 = """CLIENT INFORMATION:
ACME FENESTRATION
123 Market Street
Springfield, Illinois  62701
TEST LABORATORY:
Test Lab Co
500 Lab Road
Testville, Ohio 44000
PROJECT SUMMARY:
TEST SPECIMEN DESCRIPTION:
PRODUCT SIZES:
Overall size
920
36
1520
60
"""

NAFS_RESULTS = """TEST RESULTS: The temperature during testing was 73 F.
Test Specimen #1
OPERATING FORCE: (per ASTM E 2068)
Initiate motion
AIR LEAKAGE TESTING: (per ASTM E 283)
Infiltration @ 75 Pa (1.57 psf)
WATER PENETRATION TESTING: (ASTM E 547)
290 Pa (6.05 psf)
Pass
No Leakage
UNIFORM LOAD TESTING: (per ASTM E 330)
FORCED ENTRY RESISTANCE
Pass
No Entry
Test Specimen #2
UNIFORM LOAD TESTING: (per ASTM E 330)
Design Pressure Test
General Notes: All testing performed per reference methods.
The operating force results listed above represent the maximum.
The specimen meets air leakage resistance.
"""


def _nafs_doc():
    return Document(
        source="synthetic_nafs.pdf",
        pages=[
            Page(1, NAFS_PAGE1),
            Page(2, NAFS_PAGE2),
            Page(3, NAFS_RESULTS),
        ],
    )


TAS_PAGE = """ARCADIA TEST
Report No.: TS1234.01-200-10-R0
TAS 201, TAS 202, and TAS 203 testing on the specimen
SECTION 3 TEST METHODS
TAS 201-94, Impact Test Procedures
TAS 202-94, Criteria for Testing
TAS 203-94, Cyclic Wind Pressure
SUMMARY OF TEST RESULTS
DESIGN PRESSURE
+50 / -50 psf
REPORT ISSUED TO
ACME FENESTRATION
123 Market Street
Springfield, IL 62701
Product Type: Fixed Window
Series/Model: Demo Triple Fixed Window - 100" X
100"
Product Size:
Overall size
2540
100
2540
100
Series/Model: Demo Triple Fixed Window - 100" X 200"
Frame Construction:
MEMBER
MATERIAL
DESCRIPTION
Head
Aluminum
Extruded
Jambs
Aluminum
Extruded
Sill
Aluminum
Extruded
JOINERY TYPE
Glazing:
GLASS TYPE
OVERALL
THICKNESS
GLASS MAKEUP
GLAZING METHOD
Laminated
Glass
1.34"
1/4" Tempered
1/2 Air Gap
1/4" heat-strengthened glass
Wet glazed interior and
exterior
LOCATION
QUANTITY
DAYLIGHT OPENING
GLASS BITE
inches
Left Panel
1
68-1/2 X 116-1/16
1/2"
Drainage: No Drainage was utilized
Hardware: No Hardware was utilized
SECTION 9 CONCLUSIONS
the specimen satisfies the large missile requirements of TAS 201 and
met the requirements of Section 1626. No signs of failure were observed
in any area of the test specimen during the TAS 202 testing and the
specimens satisfy the cyclic load requirements of TAS 203.
"""


def _tas_doc():
    return Document(source="synthetic_tas.pdf", pages=[Page(1, TAS_PAGE)])


# --------------------------------------------------------------------------- #
# Tests
# --------------------------------------------------------------------------- #


def test_detect_family():
    assert detect_family(NAFS_PAGE1) == FAMILY_NAFS
    assert detect_family(TAS_PAGE) == FAMILY_TAS
    assert detect_family("just some unrelated text") == parsing.FAMILY_GENERIC


def test_labeled_longest_recovers_wrapped_value():
    # The full one-line value must win over the wrapped/truncated one.
    assert labeled_longest(TAS_PAGE, r"Series\s*/?\s*Model") == \
        'Demo Triple Fixed Window - 100" X 200"'


def test_tas_identity_and_standards_dedup():
    report = analyze(_tas_doc())
    assert report.report_number == "TS1234.01-200-10-R0"
    assert report.revision == "R0"
    assert report.standard_family == FAMILY_TAS
    # Bare "TAS 201/202/203" must be dropped in favour of the versioned names.
    assert report.test_standards == ["TAS 201-94", "TAS 202-94", "TAS 203-94"]
    assert report.product_type == "Fixed Window"
    assert "Pass" in report.overall_result


def test_tas_frame_glazing_and_dimensions():
    report = analyze(_tas_doc())
    spec = report.specimens[0]
    assert spec.overall_size == "100 in x 100 in"
    members = spec.construction.frame
    assert [m.member for m in members] == ["Head", "Jambs", "Sill"]
    assert all(m.material == "Aluminum" and m.detail == "Extruded" for m in members)
    g = spec.construction.glazing
    assert g.glass_type == "Laminated Glass"
    assert g.overall_thickness == '1.34"'
    assert g.method == "Wet glazed interior and exterior"
    assert g.bite == '1/2"'
    assert spec.daylight_opening == "68-1/2 X 116-1/16"
    # "No Hardware was utilized" -> no hardware items.
    assert spec.construction.hardware == []


def test_tas_results_pass_flags():
    report = analyze(_tas_doc())
    results = report.specimens[0].results
    assert results.get("Large Missile Impact (TAS 201)") == "Pass"
    assert results.get("Cyclic Wind Pressure (TAS 203)") == "Pass"
    assert results.get("Structural (TAS 202)") == "Pass"


def test_nafs_specimens_and_see_specimen_resolution():
    report = analyze(_nafs_doc())
    assert report.standard_family == FAMILY_NAFS
    assert report.client == "ACME FENESTRATION"
    assert report.laboratory == "Test Lab Co"
    assert len(report.specimens) == 2
    s1, s2 = report.specimens
    assert (s1.specimen_id, s1.label, s1.model) == ("1", "New Construction Frame", "300-SH")
    assert s1.product_designator == "Class R - PG40 920 x 1520 (36 x 60)-H"
    assert s1.design_pressure == "+/-1920 Pa (+/-40.10 psf)"
    assert s2.design_pressure == "+/-1440 Pa (+/-30.08 psf)"
    # "See Specimen #1" must be resolved to specimen 1's value.
    assert s2.air_infiltration == s1.air_infiltration == "0.5 L/s/m2 (0.10 cfm/ft2)"
    assert s2.water_penetration == "290 Pa (6.05 psf)"
    # Shared dimensions apply to both specimens.
    assert s1.overall_size == "36 in x 60 in"


def test_nafs_results_attribution_excludes_general_notes():
    report = analyze(_nafs_doc())
    s1, s2 = report.specimens
    assert s1.results.get("Forced Entry Resistance") == "Pass - No entry"
    assert s1.results.get("Water Penetration") == "Pass - No leakage"
    # Specimen 2 only retested uniform load; the General Notes mentioning
    # "operating force" and "air leakage" must not leak into specimen 2.
    assert set(s2.results) == {"Uniform Load (structural)"}


def test_table_helpers():
    assert _is_header_row(["Frame Member", "Material", "Detail"])
    assert not _is_header_row(["Head", "PVC", "Extruded"])
    assert _row_to_three(["a", "b", "c"]) == ("a", "b", "c")
    assert _row_to_three(["a", "b"]) == ("a", "", "b")
    assert _row_to_three(["a", "b", "c", "d"]) == ("a", "b", "c d")

    frame_table = [
        ["Frame Member", "Material", "Detail"],
        ["Head, sill and jambs", "PVC", "Extruded"],
        ["Corner construction", "--", "Miter-cut and welded"],
    ]
    frame = _frame_from_table(frame_table)
    assert [m.member for m in frame] == ["Head, sill and jambs", "Corner construction"]
    assert frame[0].material == "PVC"

    hw_table = [
        ["Description", "Quantity", "Location"],
        ["Metal keeper", "1", "Jamb"],
    ]
    hardware = _hardware_from_table(hw_table)
    assert len(hardware) == 1
    assert hardware[0].description == "Metal keeper" and hardware[0].quantity == "1"

    glaz_table = [
        ["Description", "Detail"],
        ["Glass Type", '1" IG'],
        ["Glazing Method", "Set from the exterior"],
        ["Glazing Bite", '1/2"'],
    ]
    g = _glazing_from_table(glaz_table)
    assert g.glass_type == '1" IG'
    assert g.method == "Set from the exterior"
    assert g.bite == '1/2"'


def test_finalize_raw_strips_boilerplate():
    boiler = {"TEST REPORT FOR ARCADIA INC.", "25800 Commercentre Drive"}
    raw = (
        "No Hardware was utilized\n"
        "25800 Commercentre Drive\n"
        "TEST REPORT FOR ARCADIA INC.\n"
        "Page 6 of 40\n"
        "Version: 06/21/24\n"
    )
    assert _finalize_raw(raw, boiler) == "No Hardware was utilized"


def test_flatten_and_csv_columns():
    report = analyze(_nafs_doc())
    rows = report.flatten()
    assert len(rows) == 2  # one row per specimen
    for row in rows:
        # Every row key must be a known CSV column (DictWriter would drop others).
        assert set(row).issubset(set(CSV_COLUMNS))
    assert rows[0]["report_number"] == "99001.01-100-50-R1"
    assert rows[1]["specimen_model"] == "350-SH"


def test_empty_report_still_flattens_to_one_row():
    rows = Report(source_file="x.pdf").flatten()
    assert len(rows) == 1


def test_to_dict_prunes_empty_fields():
    d = Report(source_file="x.pdf", report_number="R1").to_dict()
    assert d["report_number"] == "R1"
    assert "issue_date" not in d  # empty strings pruned


def test_xlsx_export_routes_and_maps():
    import os
    import tempfile

    import openpyxl
    from report_analyzer.xlsx_export import is_door, write_xlsx, _dp

    window = analyze(_tas_doc())     # "Fixed Window" -> FX sheet
    door = analyze(_nafs_doc())      # made a door below for routing

    assert not is_door(window)
    # Make the NAFS fixture a door for routing purposes.
    door.product_type = "PVC Patio Door, Type XO"
    assert is_door(door)

    path = os.path.join(tempfile.mkdtemp(), "summary.xlsx")
    fx, sd = write_xlsx([window, door], path)
    assert (fx, sd) == (1, 2)

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["FX Temp.", "SD Temp."]

    fxs = wb["FX Temp."]
    assert [c.value for c in fxs[1]][:6] == [
        "Test Report #", "Report Date", "Specimen #", "DP (psf)", "Test Standard", "Frame Size",
    ]
    assert fxs["A2"].value == "TS1234.01-200-10-R0"
    assert fxs["C2"].value.startswith("DEMO TRIPLE FIXED WINDOW")
    assert fxs["D2"].value == "+50 / -50 psf"
    assert fxs["F2"].value == "100 x 100"
    assert fxs["G2"].value == "68-1/2 X 116-1/16"

    sds = wb["SD Temp."]
    assert sds.max_row == 3  # header + 2 specimens
    assert (sds["C2"].value, sds["C3"].value) == ("1", "2")
    assert "New Construction Frame" in sds["L2"].value
    # Compact psf for a Pa/psf design pressure.
    assert _dp(door.specimens[0]) == "+40.10 / -40.10 psf"


def test_xlsx_date_prefers_revised():
    from report_analyzer.xlsx_export import _parse_date
    import datetime
    r = Report(report_date="12/16/2025", revised_date="01/19/2026")
    assert _parse_date(r) == datetime.date(2026, 1, 19)


# AI-extraction JSON -> Report mapping (no API call; tests the pure mapping).
AI_SAMPLE = {
    "report_number": "QCT25-7572.16",
    "standard_family": "AAMA/WDMA/CSA 101/I.S.2/A440 (NAFS)",
    "test_standards": ["AAMA/WDMA/CSA 101/I.S.2/A440-22", "AAMA 450-20"],
    "laboratory": "Quast Consulting and Testing, Inc.",
    "client": "Quaker Windows and Doors",
    "product_type": "C200 Fixed Transom",
    "series_model": "C-Mull Sidelite",
    "product_category": "window",
    "report_date": "01/21/2026",
    "test_dates": "08/25/2025 - 08/26/2025",
    "overall_result": "Pass",
    "specimens": [
        {
            "specimen_id": "1",
            "label": "C-Mull Sidelite",
            "product_designator": "Class LC - PG50: 2438 x 3658 mm (~96 x 144 in) - Type SLT",
            "design_pressure": "+50.1 / -50.1 psf",
            "air_infiltration": "0.0 cfm/ft2 @ 1.57 psf",
            "water_penetration": "No Penetration @ 7.52 psf",
            "overall_size": "96 x 144 in",
            "daylight_opening": "44.13 x 140.13 in",
            "frame": [
                {"member": "Frame", "material": "Aluminum",
                 "detail": "Thermally broken, mitered, siliconed, two corner keys"},
            ],
            "glazing": {
                "glass_type": "1\" IG",
                "makeup": "1/4\" tempered\n1/2\" air space\n1/4\" tempered",
                "method": "Set from exterior against silicone",
                "bite": "3/8\"",
            },
            "hardware": [],
            "results": [
                {"name": "Air Infiltration", "value": "PASS"},
                {"name": "Uniform Load Structural", "value": "PASS"},
                {"name": "Forced Entry Resistance", "value": "Grade 10, PASS"},
            ],
        }
    ],
}


def test_ai_mapping_builds_report():
    from report_analyzer.ai_extract import _to_report

    report = _to_report(AI_SAMPLE, source="qct.pdf")
    assert report.report_number == "QCT25-7572.16"
    assert report.laboratory == "Quast Consulting and Testing, Inc."
    assert report.product_category == "window"
    assert report.test_standards == ["AAMA/WDMA/CSA 101/I.S.2/A440-22", "AAMA 450-20"]
    assert len(report.specimens) == 1
    spec = report.specimens[0]
    assert spec.design_pressure == "+50.1 / -50.1 psf"
    assert spec.overall_size == "96 x 144 in"
    assert spec.construction.glazing.bite == '3/8"'
    assert spec.construction.frame[0].material == "Aluminum"
    assert spec.results["Forced Entry Resistance"] == "Grade 10, PASS"


def test_ai_mapping_routes_by_category():
    from report_analyzer.ai_extract import _to_report
    from report_analyzer.xlsx_export import is_door, split_rows

    window = _to_report(AI_SAMPLE, source="qct.pdf")
    door = _to_report({**AI_SAMPLE, "product_category": "door",
                       "product_type": "Sliding Patio Door"}, source="d.pdf")
    assert not is_door(window)
    assert is_door(door)
    fx, sd = split_rows([window, door])
    assert len(fx) == 1 and len(sd) == 1
    assert fx[0][0] == "QCT25-7572.16"  # report number in column A


def test_llm_parse_json_tolerates_fences_and_prose():
    from report_analyzer.llm_extract import _parse_json

    assert _parse_json('{"a": 1}') == {"a": 1}
    assert _parse_json('```json\n{"a": 1}\n```') == {"a": 1}
    assert _parse_json('Here you go:\n{"a": 1, "b": "x"}\nThanks!') == {"a": 1, "b": "x"}


def test_llm_resolve_provider_defaults_and_key():
    import os
    from report_analyzer.llm_extract import _resolve

    # Provider defaults fill in base_url + model; no key -> api_key is None.
    base, model, key = _resolve("gemini", None, None, None)
    assert "generativelanguage.googleapis.com" in base
    assert model == "gemini-2.5-flash"

    # An env key for the provider is picked up.
    os.environ["GROQ_API_KEY"] = "test-key-123"
    try:
        base, model, key = _resolve("groq", None, None, None)
        assert key == "test-key-123"
        assert "groq.com" in base
    finally:
        del os.environ["GROQ_API_KEY"]

    # Ollama needs no real key.
    base, model, key = _resolve("ollama", None, None, None)
    assert key == "ollama" and "11434" in base


def test_load_dotenv_sets_and_preserves():
    import os
    import tempfile
    from report_analyzer.cli import _load_dotenv

    path = os.path.join(tempfile.mkdtemp(), ".env")
    with open(path, "w") as f:
        f.write("# a comment\nRA_TEST_KEY = hello123 \nRA_EXISTING='fromfile'\n")
    os.environ.pop("RA_TEST_KEY", None)
    os.environ["RA_EXISTING"] = "preset"
    try:
        _load_dotenv(path)
        assert os.environ["RA_TEST_KEY"] == "hello123"          # loaded + trimmed
        assert os.environ["RA_EXISTING"] == "preset"            # existing not overridden
    finally:
        os.environ.pop("RA_TEST_KEY", None)
        os.environ.pop("RA_EXISTING", None)


def test_llm_skeleton_matches_schema_shape():
    from report_analyzer.ai_extract import SCHEMA
    from report_analyzer.llm_extract import _skeleton

    sk = _skeleton(SCHEMA)
    assert sk["report_number"] == ""
    assert isinstance(sk["specimens"], list) and isinstance(sk["specimens"][0], dict)
    assert sk["specimens"][0]["design_pressure"] == ""
    assert isinstance(sk["specimens"][0]["frame"], list)


# --------------------------------------------------------------------------- #
# Script runner (no pytest required)
# --------------------------------------------------------------------------- #

if __name__ == "__main__":
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failed = 0
    for fn in tests:
        try:
            fn()
            print(f"PASS {fn.__name__}")
        except AssertionError as exc:
            failed += 1
            print(f"FAIL {fn.__name__}: {exc}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"ERROR {fn.__name__}: {type(exc).__name__}: {exc}")
    print(f"\n{len(tests) - failed}/{len(tests)} passed")
    sys.exit(1 if failed else 0)
